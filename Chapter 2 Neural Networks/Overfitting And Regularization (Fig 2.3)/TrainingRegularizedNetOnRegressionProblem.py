import torch
# Test Effect of different training methods on the conservatism of theta_m (Combettes)

import numpy as np
import matlab
import matplotlib.pyplot as plt
import os
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.autograd import Variable
import torch.optim as optim
import time
from datetime import datetime
from scipy.io import savemat
from openpyxl import workbook, load_workbook

# # load from Excel
# wb = load_workbook('Daten_trainedNetwork_2DToyExample_1x20.xlsx')
# # get active worksheet in workbook
# ws = wb.active
# # latest Round is safed in (1,1)
# Round = ws.cell(1, 1).value
# Round = Round + 1

INPUT_SIZE = 1
HIDDEN_SIZE1 = 20
HIDDEN_SIZE2 = 20
HIDDEN_SIZE3 = 20
OUTPUT_SIZE = 1

# hyperperparameters
lr = 0.1 # learning rate #prior 0.1
lmbd_l1 = 0.01  # L1 penalty parameter #prior:0.0005
lmbd_l2 = 0.01 # L2 penalty parameter

class simpleNet(nn.Module):
    def __init__(self):
        super(simpleNet, self).__init__()
        self.lin1 = nn.Linear(INPUT_SIZE, HIDDEN_SIZE1)
        self.lin2 = nn.Linear(HIDDEN_SIZE1, HIDDEN_SIZE2)
        self.lin3 = nn.Linear(HIDDEN_SIZE2, HIDDEN_SIZE3)
        self.lin4 = nn.Linear(HIDDEN_SIZE3, OUTPUT_SIZE)

    def forward(self, x):
        x = torch.tanh(self.lin1(x))
        x = torch.tanh(self.lin2(x))
        x = torch.tanh(self.lin3(x))
        x = self.lin4(x)
        return x

    def num_flat_features(self, x):
        size = x.size()[1:]
        num = 1
        for i in size:
            num *= i
        return num

    def l2_reg(self, lmbd):         # Frage: Warum werden die bias Vektoren im penalty term berücksichtigt? Warum wird noch mit 0.5 mulitpliziert? Damit einfach anderes lambda?
        reg_loss = None
        weight_p, bias_p = [], []
        for name, p in self.named_parameters():
            if 'bias' in name:
                bias_p += [p]
            else:
                weight_p += [p]

        for param in weight_p:
            if reg_loss is None:
                reg_loss = 0.5 * torch.sum(param**2)
            else:
                reg_loss = reg_loss + 0.5 * param.norm(2)**2
        return lmbd * reg_loss

    def l1_reg(self, lmbd):        # Auch hier bias Vektoren mit einbezogen
        reg_loss = None
        weight_p, bias_p = [], []
        for name, p in self.named_parameters():
            if 'bias' in name:
                bias_p += [p]
            else:
                weight_p += [p]

        for param in weight_p:
            if reg_loss is None:
                reg_loss = 0.5 * torch.sum(abs(param))
            else:
                reg_loss = reg_loss + 0.5 * torch.sum(abs(param))
        return lmbd * reg_loss

    def extract_weights(self):
        weights = []
        biases = []
        for param_tensor in self.state_dict():
            tensor = self.state_dict()[param_tensor].detach().numpy().astype(np.float64)

            if 'weight' in param_tensor:
                weights.append(tensor)
            if 'bias' in param_tensor:
                biases.append(tensor)
        return weights, biases

    def train(self, lmbd_l1=None, lmbd_l2=None, parameters=None):
        out = self(training_input)
        criterion = nn.MSELoss()
        loss = criterion(out, training_output)
        loss_prev = 0
        loss_prevprev = 0

        for i in range(30000):
            out = self(training_input)

            loss_prev = loss
            loss = criterion(out, training_output)

            if lmbd_l1 is not None:
                loss += self.l1_reg(lmbd_l1)

            if lmbd_l2 is not None:
                loss += self.l2_reg(lmbd_l2)

            if np.mod(i+1, 1000) == 0:

                MSELoss = criterion(out, training_output)
                print('Train Epoch: {}; Loss: {:.6f}; MSE-Loss: {:.6f}'.format(
                    i, loss.item(), MSELoss))

            self.zero_grad()
            loss.backward()

            optimizer = optim.SGD(self.parameters(), lr=lr)
            optimizer.step()


# data for regression problem
training_input = []
training_output = []
x = - 3
counter1 = 0
while x <= 3:
    y = np.power(x,2) + 1.8*np.random.rand()-0.9
    training_input.append([x])
    training_output.append([y])
    x = x + 0.1
    counter1 = counter1 + 1
print(training_input)

plt.scatter(training_input, training_output)
plt.title('training data')
plt.show()

training_input = torch.Tensor(training_input)
training_output = torch.Tensor(training_output)

# Vanilla Network
net_vanilla= simpleNet()
print("Beginnning nominal NN training")
t = time.time()

# train until decent value found
criterion = nn.MSELoss()
loss = 1
out = net_vanilla(training_input)
while loss > 0.2:
    net_vanilla.train()
    network_output = net_vanilla(training_input)
    loss = criterion(network_output, training_output)
timeNom = time.time() - t
print("Nominal Training Complete after {} seconds".format(timeNom))
weights, biases = net_vanilla.extract_weights()
name_str = 'NomNetwork_NoisyParabula.pth'
torch.save(net_vanilla.state_dict(), name_str)

# L2 Regularized network
net_L1 = simpleNet()
print("Beginnning L1 training")
t = time.time()

criterion = nn.MSELoss()
loss = 1
while loss > 0.2:
    net_L1.train(lmbd_l2= lmbd_l2)
    network_output = net_vanilla(training_input)
    loss = criterion(network_output, training_output)
timeNom = time.time() - t
print("L1 Training Complete after {} seconds".format(timeNom))
weights, biases = net_vanilla.extract_weights()
name_str = 'L1RegularizedNetwork_NoisyParabula.pth'
torch.save(net_vanilla.state_dict(), name_str)

plotting_input = []
x = - 3
counter2 = 0
x_plotting_array = []

while x <= 3:
    x_plotting_array.append(x)
    y = np.power(x,2) + 1.4*np.random.rand()-0.7
    plotting_input.append([x])
    x = x + 0.01
    counter2 = counter2 + 1
plotting_input = torch.Tensor(plotting_input)

# Plot resulting functions
out_Nom = net_vanilla(plotting_input).detach()
#print(out_Nom)
out_L1 = net_L1(plotting_input).detach()
plt.scatter(training_input, training_output)
plt.scatter(plotting_input, out_Nom)
plt.scatter(plotting_input, out_L1 )
plt.title('Effect of regularization')
plt.show()

#print(out_Nom.size)
# convert array of individual arrays to a single array
# save data for matlab
out_Nom_singleArray = []
out_L1_singleArray = []
training_input_singleArray = []
training_output_singleArray = []

for i in range(counter2):
    out_Nom_singleArray.append(out_Nom[i][0])
    out_L1_singleArray.append(out_L1[i][0])

for j in range(counter1):
    training_input_singleArray.append(training_input[j][0])
    training_output_singleArray.append(training_output[j][0])

savemat("out_Nom_singleArray.mat", {"out_Nom_singleArray": out_Nom_singleArray}, do_compression=False)
savemat("out_L1_singleArray.mat", {"out_L1_singleArray": out_L1_singleArray}, do_compression=False)
savemat("training_input_singleArray.mat", {"training_input_singleArray": training_input_singleArray}, do_compression=False)
savemat("training_output_singleArray.mat", {"training_output_singleArray": training_output_singleArray}, do_compression=False)
savemat("x_plotting_array.mat", {"x_plotting_array": x_plotting_array}, do_compression=False)