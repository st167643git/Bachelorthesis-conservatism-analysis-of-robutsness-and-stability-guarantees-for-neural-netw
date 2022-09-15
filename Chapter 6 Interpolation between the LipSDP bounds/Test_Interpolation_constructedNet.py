from MILP_LipSDP_Interpolation import *
import numpy as np
from LipSDP import lipSDP
from ConstructedNet import *
from openpyxl import workbook, load_workbook

n_layers = 20
n_Neurons_per_layer = 20
factor = 1
n_DOFs = 2
n_iter = 100

# determine net dim
net_dim = []
net_dim.append(n_Neurons_per_layer)
for layer in range(n_layers):
    net_dim.append(n_Neurons_per_layer)
net_dim.append(n_Neurons_per_layer)

weights = constructedNet1(n_layers, n_Neurons_per_layer, factor)
#weights = constructedNet2(n_layers, n_Neurons_per_layer, factor)

# load excel workbook
wb = load_workbook('LipSDPInterpolateConstructedNet1.xlsx')
ws = wb.active

# compute and save results, computation times are averaged due to high variance in computation time when running multiple times

time_total_solver_MILP_array = []
time_overhead_MILP_array = []
time_solver_SDP_array = []
time_overhead_SDP_array = []

for i in range(n_iter+20):
    L_LipSDP_Interpolate_MILP, time_total_solver_MILP, time_overhead_MILP, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_MILP(weights, n_DOFs, 10**3) #int(n_Neurons_per_layer/2)
    # for some reason solver get significantly faster after solving a few problems, -> dont save first time values
    if i >= 20:
        time_total_solver_MILP_array.append(time_total_solver_MILP)
        time_overhead_MILP_array.append(time_overhead_MILP)
        time_solver_SDP_array. append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
ws.cell(2, 2).value = L_LipSDP_Interpolate_MILP
ws.cell(3, 2).value = np.mean(time_total_solver_MILP_array)
ws.cell(4, 2).value = np.mean(time_overhead_MILP_array)
ws.cell(5, 2).value = np.mean(time_solver_SDP_array)
ws.cell(6, 2).value = np.mean(time_overhead_SDP_array)

time_overhead_Heuristik_array = []
time_solver_SDP_array = []
time_overhead_SDP_array = []
for i in range(n_iter):
    L_LipSDP_Interpolate_Heuristik, time_overhead_Heuristik, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_Heuristik(weights, n_DOFs)
    time_overhead_Heuristik_array.append(time_overhead_Heuristik)
    time_solver_SDP_array.append(time_solver_SDP)
    time_overhead_SDP_array.append(time_overhead_SDP)
ws.cell(2, 3).value = L_LipSDP_Interpolate_Heuristik
ws.cell(4, 3).value = np.mean(time_overhead_Heuristik_array)
ws.cell(5, 3).value = np.mean(time_solver_SDP_array)
ws.cell(6, 3).value = np.mean(time_overhead_SDP_array)


for j in range(4):
    n_DOFs_rand = j + 2
    L_LipSDP_Interpolate_Random_array = []
    time_solver_SDP_array = []
    time_overhead_SDP_array = []
    for i in range(n_iter):
        L_LipSDP_Interpolate_Random, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_Random(weights, n_DOFs_rand)
        L_LipSDP_Interpolate_Random_array.append(L_LipSDP_Interpolate_Random)
        time_solver_SDP_array.append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
    ws.cell(2, j + 4).value = np.mean(L_LipSDP_Interpolate_Random_array)
    ws.cell(5, j + 4).value = np.mean(time_solver_SDP_array)
    ws.cell(6, j + 4).value = np.mean(time_overhead_SDP_array)

time_solver_SDP_array = []
time_overhead_SDP_array = []
for i in range(n_iter):
    L_LipSDP_Neuron, time_solver_SDP, time_overhead_SDP = lipSDP(weights, net_dim, 'neuron')
    time_solver_SDP_array.append(time_solver_SDP)
    time_overhead_SDP_array.append(time_overhead_SDP)
ws.cell(2, 8).value = L_LipSDP_Neuron
ws.cell(5, 8).value = np.mean(time_solver_SDP_array)
ws.cell(6, 8).value = np.mean(time_overhead_SDP_array)

time_solver_SDP_array = []
time_overhead_SDP_array = []
for i in range(n_iter):
    L_LipSDP_Layer, time_solver_SDP, time_overhead_SDP = lipSDP(weights, net_dim, 'layer')
    time_solver_SDP_array.append(time_solver_SDP)
    time_overhead_SDP_array.append(time_overhead_SDP)
ws.cell(2, 9).value = L_LipSDP_Layer
ws.cell(5, 9).value = np.mean(time_solver_SDP_array)
ws.cell(6, 9).value = np.mean(time_overhead_SDP_array)

# save workbook
wb.save('LipSDPInterpolateConstructedNet1.xlsx')

print("LipSDP_Interpolate_MILP", L_LipSDP_Interpolate_MILP)
print("LipSDP_Interpolate_Heuristik", L_LipSDP_Interpolate_Heuristik)
print("LipSDP_Interpolate_Random", L_LipSDP_Interpolate_Random)
print("LipSDP_Neuron", L_LipSDP_Neuron)
print("LipSDP_Layer", L_LipSDP_Layer)

