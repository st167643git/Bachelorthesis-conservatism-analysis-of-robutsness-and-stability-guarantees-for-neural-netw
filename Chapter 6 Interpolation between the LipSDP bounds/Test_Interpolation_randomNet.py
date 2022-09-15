from MILP_LipSDP_Interpolation import *
import numpy as np
import matlab
import matlab.engine
from LipSDP import lipSDP
from openpyxl import workbook,load_workbook

eng = matlab.engine.start_matlab()

# W_prior =  np.random.uniform(-1, 1, size=(10, 10)) #[[ 1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1, 1], [1, 1, 1, 1]]
# #W_prior = np.matrix(W_prior)
# W = np.random.uniform(-1, 1, size=(10, 10))#[[1, 1, 2, 2], [1, 1, 2, 2], [1, 1, 2, 2], [1, 1, 2, 2]]
# #W = np.matrix(W)
# x = interpol_Random(W, W_prior, 9)
# print(x)

n_layers = 6
n_Neurons_per_layer = 20
n_DOFs = 5
n_iter = 1

net_dim = []
net_dim.append(n_Neurons_per_layer)
for layer in range(n_layers):
    net_dim.append(n_Neurons_per_layer)
net_dim.append(n_Neurons_per_layer)

# load excel workbook
wb = load_workbook('LipSDPInterpolateRandomNet_fertig.xlsx')
ws = wb.active

# compute and save results, computation times are averaged due to high variance in computation time when running multiple times
#multiple rounds with random matrices
for n_round in range(50):

    weights = []
    for W in range(n_layers + 1):
        W_mat = np.random.uniform(-1, 1, size=(n_Neurons_per_layer, n_Neurons_per_layer))
        weights.append(W_mat)

    time_total_solver_MILP_array = []
    time_overhead_MILP_array = []
    time_solver_SDP_array = []
    time_overhead_SDP_array = []

    for i in range(n_iter + 20):
        L_LipSDP_Interpolate_MILP, time_total_solver_MILP, time_overhead_MILP, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_MILP(
            weights, n_DOFs, 10 ** 3)  # int(n_Neurons_per_layer/2)
        # for some reason solver get significantly faster after solving a few problems, -> dont save first time values
        if i >= 20:
            time_total_solver_MILP_array.append(time_total_solver_MILP)
            time_overhead_MILP_array.append(time_overhead_MILP)
            time_solver_SDP_array.append(time_solver_SDP)
            time_overhead_SDP_array.append(time_overhead_SDP)
    ws.cell(2 + n_round * 5, 2).value = L_LipSDP_Interpolate_MILP
    ws.cell(3 + n_round * 5, 2).value = np.mean(time_total_solver_MILP_array)
    ws.cell(4 + n_round * 5, 2).value = np.mean(time_overhead_MILP_array)
    ws.cell(5 + n_round * 5, 2).value = np.mean(time_solver_SDP_array)
    ws.cell(6 + n_round * 5, 2).value = np.mean(time_overhead_SDP_array)

    time_overhead_Heuristik_array = []
    time_solver_SDP_array = []
    time_overhead_SDP_array = []
    for i in range(n_iter):
        L_LipSDP_Interpolate_Heuristik, time_overhead_Heuristik, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_Heuristik(
            weights, n_DOFs)
        time_overhead_Heuristik_array.append(time_overhead_Heuristik)
        time_solver_SDP_array.append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
    ws.cell(2 + n_round * 5, 3).value = L_LipSDP_Interpolate_Heuristik
    ws.cell(4 + n_round * 5, 3).value = np.mean(time_overhead_Heuristik_array)
    ws.cell(5 + n_round * 5, 3).value = np.mean(time_solver_SDP_array)
    ws.cell(6 + n_round * 5, 3).value = np.mean(time_overhead_SDP_array)

    L_LipSDP_Interpolate_Random_array = []
    time_solver_SDP_array = []
    time_overhead_SDP_array = []
    # initialize lowest bound by random networks with a large number
    L_min = 10**10
    for i in range(100):
        L_LipSDP_Interpolate_Random, time_solver_SDP, time_overhead_SDP = LipSDP_Interpolate_Random(weights, n_DOFs)
        L_LipSDP_Interpolate_Random_array.append(L_LipSDP_Interpolate_Random)
        time_solver_SDP_array.append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
        if L_LipSDP_Interpolate_Random < L_min:
            L_min = L_LipSDP_Interpolate_Random

    ws.cell(2 + n_round * 5, 4).value = np.mean(L_LipSDP_Interpolate_Random_array)
    ws.cell(5 + n_round * 5, 4).value = np.mean(time_solver_SDP_array)
    ws.cell(6 + n_round * 5, 4).value = np.mean(time_overhead_SDP_array)
    ws.cell(2 + n_round * 5, 5).value = L_min

    time_solver_SDP_array = []
    time_overhead_SDP_array = []
    for i in range(n_iter):
        L_LipSDP_Neuron, time_solver_SDP, time_overhead_SDP = lipSDP(weights, net_dim, 'neuron')
        time_solver_SDP_array.append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
    ws.cell(2 + n_round * 5, 6).value = L_LipSDP_Neuron
    ws.cell(5 + n_round * 5, 6).value = np.mean(time_solver_SDP_array)
    ws.cell(6 + n_round * 5, 6).value = np.mean(time_overhead_SDP_array)

    time_solver_SDP_array = []
    time_overhead_SDP_array = []
    for i in range(n_iter):
        L_LipSDP_Layer, time_solver_SDP, time_overhead_SDP = lipSDP(weights, net_dim, 'layer')
        time_solver_SDP_array.append(time_solver_SDP)
        time_overhead_SDP_array.append(time_overhead_SDP)
    ws.cell(2 + n_round * 5, 7).value = L_LipSDP_Layer
    ws.cell(5 + n_round * 5, 7).value = np.mean(time_solver_SDP_array)
    ws.cell(6 + n_round * 5, 7).value = np.mean(time_overhead_SDP_array)

    # save workbook after each round
    wb.save('LipSDPInterpolateRandomNet_fertig.xlsx')


print("LipSDP_Interpolate_MILP", L_LipSDP_Interpolate_MILP)
print("LipSDP_Interpolate_Heuristik", L_LipSDP_Interpolate_Heuristik)
print("LipSDP_Interpolate_Random", L_LipSDP_Interpolate_Random)
print("LipSDP_Neuron", L_LipSDP_Neuron)
print("LipSDP_Layer", L_LipSDP_Layer)