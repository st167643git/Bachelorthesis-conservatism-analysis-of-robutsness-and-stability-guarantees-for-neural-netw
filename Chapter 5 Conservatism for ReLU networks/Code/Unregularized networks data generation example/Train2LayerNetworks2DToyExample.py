import torch
# Test Effect of different training methods on the conservatism of theta_m (Combettes)

import numpy as np
import matplotlib.pyplot as plt
import os
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.autograd import Variable
import torch.optim as optim
import time
from datetime import datetime
from scipy.io import savemat
from openpyxl import workbook, load_workbook

# load from Excel
wb = load_workbook('Daten_trainedNetwork_2DToyExample_1x20.xlsx')
# get active worksheet in workbook
ws = wb.active
# latest Round is safed in (1,1)
Round = ws.cell(1,1).value
Round = Round + 1

INPUT_SIZE = 2
HIDDEN_SIZE1 = 20
#HIDDEN_SIZE2 = 10
OUTPUT_SIZE = 3

# hyperperparameters
lr = 0.1 # learning rate #prior 0.1
lmbd_l1 = 0.001  # L1 penalty parameter #prior:0.0005
lmbd_l2 = 0.001 # L2 penalty parameter

class simpleNet(nn.Module):
    def __init__(self):
        super(simpleNet, self).__init__()
        self.lin1 = nn.Linear(INPUT_SIZE, HIDDEN_SIZE1)
        self.lin2 = nn.Linear(HIDDEN_SIZE1, OUTPUT_SIZE)

        self.initialize_weights()

    def forward(self, x):
        x = torch.relu(self.lin1(x))
        x = self.lin2(x)
        return x

    def num_flat_features(self, x):
        size = x.size()[1:]
        num = 1
        for i in size:
            num *= i
        return num

    def l2_reg(self, lmbd):         # Frage: Warum werden die bias Vektoren im penalty term berücksichtigt? Warum wird noch mit 0.5 mulitpliziert? Damit einfach anderes lambda?
        reg_loss = None
        weight_p, bias_p = [], []
        for name, p in self.named_parameters():
            if 'bias' in name:
                bias_p += [p]
            else:
                weight_p += [p]

        for param in weight_p:
            if reg_loss is None:
                reg_loss = 0.5 * torch.sum(param**2)
            else:
                reg_loss = reg_loss + 0.5 * param.norm(2)**2
        return lmbd * reg_loss

    def l1_reg(self, lmbd):        # Auch hier bias Vektoren mit einbezogen
        reg_loss = None
        weight_p, bias_p = [], []
        for name, p in self.named_parameters():
            if 'bias' in name:
                bias_p += [p]
            else:
                weight_p += [p]

        for param in weight_p:
            if reg_loss is None:
                reg_loss = 0.5 * torch.sum(abs(param))
            else:
                reg_loss = reg_loss + 0.5 * torch.sum(abs(param))
        return lmbd * reg_loss

    def extract_weights(self):
        weights = []
        biases = []
        for param_tensor in self.state_dict():
            tensor = self.state_dict()[param_tensor].detach().numpy().astype(np.float64)

            if 'weight' in param_tensor:
                weights.append(tensor)
            if 'bias' in param_tensor:
                biases.append(tensor)
        return weights, biases

    def initialize_weights(self):
        for m in self.modules():
            if isinstance(m,nn.Linear):
                nn.init.uniform_(m.weight, -1,1)
                nn.init.uniform_(m.bias, -1, 1)

    def train(self, lmbd_l1=None, lmbd_l2=None, parameters=None):
        out = self(input)
        criterion = nn.CrossEntropyLoss()
        loss = criterion(out, target_cross)
        loss_prev = 0
        loss_prevprev = 0

        for i in range(10000):
            out = self(input)

            loss_prev = loss
            loss = criterion(out, target_cross)

            if lmbd_l1 is not None:
                loss += self.l1_reg(lmbd_l1)

            if lmbd_l2 is not None:
                loss += self.l2_reg(lmbd_l2)

            if np.mod(i+1, 1000) == 0:

                crossEntropyLoss = criterion(out, target_cross)
                print('Train Epoch: {}; Loss: {:.6f}; CE-Loss: {:.6f}'.format(
                    i, loss.item(), crossEntropyLoss))

            self.zero_grad()
            loss.backward()

            optimizer = optim.SGD(self.parameters(), lr=lr)
            #optimizer = optim.Adagrad(self.parameters(), lr=lr)
            #optimizer = optim.Adam(self.parameters(), lr=lr)
            optimizer.step()

            while abs(loss_prevprev - loss.item()) >= 0.01:     # Warum dieser Teil? hilft das vielleicht?
                out = self(input)

                loss_prevprev = loss_prev
                loss_prev = loss
                criterion = nn.CrossEntropyLoss()
                loss = criterion(out, target_cross)

                if lmbd_l1 is not None:
                    loss += self.l1_reg(lmbd_l1)

                if lmbd_l2 is not None:
                    loss += self.l2_reg(lmbd_l2)

                self.zero_grad()
                loss.backward()

                optimizer = optim.SGD(self.parameters(), lr=lr)
                # optimizer = optim.Adagrad(self.parameters(), lr=lr)
                # optimizer = optim.Adam(self.parameters(), lr=lr)
                optimizer.step()


# 2d toy example data
# Create training Data
N = 10000
np.random.seed(1612111977)
x = np.random.rand(N, 1)
y = np.random.rand(N, 1)

# Create Input
input = torch.Tensor(np.concatenate((x, y), axis=1))
print(input)
# Create Target
target_cross = Variable(torch.zeros(N, dtype=torch.long))
for j in range(N):
    if (x[j]-0.5)**2 + (y[j]-0.5)**2 <= 0.04:
        target_cross[j] += 2
    elif (x[j]-0.5)**2 + (y[j]-0.5)**2 <= 0.16:
        target_cross[j] += 1
    else:
        target_cross[j] += 0

# plot true classifications
x0_true = []
x1_true = []
x2_true = []
y0_true = []
y1_true = []
y2_true = []
for i in range(len(x)):
    if target_cross[i].item() == 0:
        x0_true = np.append(x0_true, x[i])
        y0_true = np.append(y0_true, y[i])
    elif target_cross[i].item() == 1:
        x1_true = np.append(x1_true, x[i])
        y1_true = np.append(y1_true, y[i])
    if target_cross[i].item() == 2:
        x2_true = np.append(x2_true, x[i])
        y2_true = np.append(y2_true, y[i])

# circle_green = plt.Circle((0.5, 0.5), 0.2, color='g', fill=False)
# circle_orange = plt.Circle((0.5, 0.5), 0.4, color='r', fill=False)
# fig, ax = plt.subplots()
# ax.add_artist(circle_green)
# ax.add_artist(circle_orange)

# plt.scatter(x0_true, y0_true)
# plt.scatter(x1_true, y1_true)
# plt.scatter(x2_true, y2_true)
# plt.title('training data')
# plt.show()


# Create Vanilla Network
net_vanilla= simpleNet()
print("Beginnning nominal NN training")
t = time.time()
optimizer = optim.SGD(net_vanilla.parameters(), lr=lr)
# train until decent value found
criterion = nn.CrossEntropyLoss()
out = net_vanilla(input)
loss = criterion(out, target_cross)
while loss > 0.2:
    net_vanilla.initialize_weights()
    net_vanilla.train(lmbd_l1= lmbd_l1)
    out = net_vanilla(input)
    loss = criterion(out, target_cross)
timeNom = time.time() - t
print("Nominal Training Complete after {} seconds".format(timeNom))
weights, biases = net_vanilla.extract_weights()
name_str = '2D_NomModel_1x20_stateDict_L1' + str(Round) + '.pth'
torch.save(net_vanilla.state_dict(), name_str)

# # # NN with L2 regularizer
# net_L2 = simpleNet()
# print("Beginnning L2 training")
# t = time.time()
# optimizer = optim.SGD(net_L2.parameters(), lr=lr)
# net_L2.train(lmbd_l2=lmbd_l2)
# timeL2 = time.time() - t
# print("L2 Training Complete after {} seconds".format(timeL2))
# weights_L2, biases_L2 = net_L2.extract_weights()
# torch.save(net_L2.state_dict(), '2D_L2Model_stateDict.pth')

# # # NN with L1 regularizer
# net_L1 = simpleNet()
# print("Beginnning L1 training")
# t = time.time()
# optimizer = optim.SGD(net_L1.parameters(), lr=lr)
# net_L1.train(lmbd_l1=lmbd_l1)
# timeL1 = time.time() - t
# print("L1 Training Complete after {} seconds".format(timeL1))
# weights_L1, biases_L1 = net_L1.extract_weights()
# torch.save(net_L1.state_dict(), '2D_L1Model_stateDict.pth')

# create more data for testing accuracy
N2 = 100000
np.random.seed(342535)
x_test = np.random.rand(N2, 1)
y_test = np.random.rand(N2, 1)

data_test = torch.Tensor(np.concatenate((x_test, y_test), axis=1))

data_test_target = Variable(torch.zeros(N2, dtype=torch.long))
for j in range(N2):
    if (x_test[j]-0.5)**2 + (y_test[j]-0.5)**2 <= 0.04:
        data_test_target[j] += 2
    elif (x_test[j]-0.5)**2 + (y_test[j]-0.5)**2 <= 0.16:
        data_test_target[j] += 1
    else:
        data_test_target[j] += 0

# Predictions on testing data
out = F.softmax(net_vanilla(data_test))
# out_L2 = F.softmax(net_L2(data_test))
# out_L1 = F.softmax(net_L1(data_test))

# Predictions on training data
out_train = F.softmax(net_vanilla(input))
# out_L2_train = F.softmax(net_L2(input))
# out_L1_train = F.softmax(net_L1(input))

# plot classifications by vanilla net
x0 = []
x1 = []
x2 = []
y0 = []
y1 = []
y2 = []
for i in range(N2):
    if np.argmax(out.detach().numpy()[i, :]) == 0:
        x0 = np.append(x0, x_test[i])
        y0 = np.append(y0, y_test[i])
    elif np.argmax(out.detach().numpy()[i, :]) == 1:
        x1 = np.append(x1, x_test[i])
        y1 = np.append(y1, y_test[i])
    if np.argmax(out.detach().numpy()[i, :]) == 2:
        x2 = np.append(x2, x_test[i])
        y2 = np.append(y2, y_test[i])

# circle_green = plt.Circle((0.5, 0.5), 0.2, color='g', fill=False)
# circle_orange = plt.Circle((0.5, 0.5), 0.4, color='r', fill=False)
# fig, ax = plt.subplots()
# ax.add_artist(circle_green)
# ax.add_artist(circle_orange)

# plt.scatter(x0, y0)
# plt.scatter(x1, y1)
# plt.scatter(x2, y2)
# plt.title('predicted classification vanilla net')
# plt.show()

# accuracy on training data
n_cor_pred = 0.0
for i in range(N):
    if np.argmax(out_train.detach().numpy()[i, :]) == target_cross[i].item():
        n_cor_pred += 1
acc_net_vanilla = n_cor_pred/N
print(f'Accuracy of vanilla network on training data: = {acc_net_vanilla:.3f}')

# n_cor_pred = 0
# for i in range(N):
#     if np.argmax(out_L2_train.detach().numpy()[i, :]) == target_cross[i].item():
#         n_cor_pred += 1
# acc_net_l2 = n_cor_pred/N
# print(f'Accuracy of l2-regularized network on training data: = {acc_net_l2:.3f}')
#
# n_cor_pred = 0
# for i in range(N):
#     if np.argmax(out_L1_train.detach().numpy()[i, :]) == target_cross[i].item():
#         n_cor_pred += 1
# acc_net_l1 = n_cor_pred/N
# print(f'Accuracy of l1-regularized on training data network: = {acc_net_l1:.3f}')

# calculate accuracy on new data
n_cor_pred = 0.0
for i in range(N2):
    if np.argmax(out.detach().numpy()[i, :]) == data_test_target[i].item():
        n_cor_pred += 1
acc_net_vanilla = n_cor_pred/N2
print(f'Accuracy of vanilla network: = {acc_net_vanilla:.3f}')

# n_cor_pred = 0
# for i in range(N2):
#     if np.argmax(out_L2.detach().numpy()[i, :]) == data_test_target[i].item():
#         n_cor_pred += 1
# acc_net_l2 = n_cor_pred/N2
# print(f'Accuracy of l2-regularized network: = {acc_net_l2:.3f}')
#
# n_cor_pred = 0
# for i in range(N2):
#     if np.argmax(out_L1.detach().numpy()[i, :]) == data_test_target[i].item():
#         n_cor_pred += 1
# acc_net_l1 = n_cor_pred/N2
# print(f'Accuracy of l1-regularized network: = {acc_net_l1:.3f}')

# write to Excel
# change value in Excel
ws.cell(Round+1, 2).value = acc_net_vanilla
ws.cell(1, 1).value = Round
wb.save('Daten_trainedNetwork_2DToyExample_1x20.xlsx')

# tests
#model = simpleNet()
#for params in model.parameters():
    #print(params)
#print("next")
#c = torch.tensor([[-1, 2, -3], [4, -5, 6]], dtype=torch.float32)
#print(c.norm(2) ** 2)
#print(torch.sum(abs(c)))

