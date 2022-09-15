import torch
import torch.nn as nn
import numpy as np
from openpyxl import workbook, load_workbook

INPUT_SIZE = 2
HIDDEN_SIZE1 = 20
#HIDDEN_SIZE2 = 10
OUTPUT_SIZE = 3


class simpleNet(nn.Module):
    def __init__(self):
        super(simpleNet, self).__init__()
        self.lin1 = nn.Linear(INPUT_SIZE, HIDDEN_SIZE1)
        self.lin2 = nn.Linear(HIDDEN_SIZE1, OUTPUT_SIZE)


    def forward(self, x):
        x = torch.relu(self.lin1(x))
        x = self.lin2(x)
        return x

    def extract_weights(self):
        weights = []
        biases = []
        for param_tensor in self.state_dict():
            tensor = self.state_dict()[param_tensor].detach().numpy().astype(np.float64)

            if 'weight' in param_tensor:
                weights.append(tensor)
            if 'bias' in param_tensor:
                biases.append(tensor)
        return weights, biases

# function to convert weights and biases into the right format for LipBaB
def create_parameter_lists(model):
    params = model.state_dict()
    weights,biases=[None],[None]

    for key, value in params.items():
        if ('weight' in key):
            weights.append(value.tolist())
        if ('bias' in key):
            biases.append(value.tolist())
    return weights, biases

# # load from Excel
# wb = load_workbook('Daten_trainedNetwork_2DToyExample_nominalNet.xlsx')
# # get active worksheet in workbook
# ws = wb.active
# # latest Round is safed in (1,1)
# Round = ws.cell(1,1).value

# # load trained networks
# net_vanilla = simpleNet()
# namestring = '2D_NomModel_stateDict' + str(Round) + '.pth'
# net_vanilla.load_state_dict(torch.load(namestring))
# net_vanilla.eval()
# net_L2 = simpleNet()
# net_L2.load_state_dict(torch.load('2D_L2Model_stateDict.pth'))
# net_L2.eval()
# net_L1 = simpleNet()
# net_L1.load_state_dict(torch.load('2D_L1Model_stateDict.pth'))
# net_L1.eval()
